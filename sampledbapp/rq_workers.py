from sampledbapp.models import *
from django.utils.timezone import now
import json
from django.core.files.storage import default_storage
import openpyxl
import logging
from  sampledbapp.utils import *
from datetime import datetime, timedelta

# Method for committing finished staging changes
def commit_staging_samples(job_id,staging_id):

    job = start_job(job_id)

    # Get the staging object and commit it.
    staging_object = Staging.objects.filter(id=staging_id,status=3)[0]

    # IDK why this is double encoded. Just leave it for now.
    json_objects = json.loads(json.loads(staging_object.json))

    for row_id,sample_fields in json_objects['objects'].items():

        # ITS NEW!
        if not is_number(row_id):

            sample_object = Sample(**sample_fields)
            sample_object.project = staging_object.project

        else:

            update_list = update_sample_model(row_id,sample_fields,staging_object.project)
            sample_object = update_list[0]

        sample_object.save()

    staging_object.status = 4
    staging_object.datetime_committed = now()
    staging_object.save()

    # Update the job status
    finish_job(job)

# Start the job in the db.
def start_job(job_id):

    # Update the job status as running.
    job = Job.objects.get(id=job_id)
    job.status = 2
    job.datetime_run = now()
    job.save()
    return job

def finish_job(job):

    # Update the job status
    job.status = 4
    job.datetime_completed = now()
    job.save()
    return job

def update_sample_model(sample_pk,sample_fields,project):

    sample_object = Sample.objects.filter(pk=sample_pk,project=project)[0]

    edited_sample_fields = []

    # study_title
    if sample_fields['study_title'] == '' or not sample_fields['study_title']:
        sample_object.study_title = None
    else:
        if sample_object.study_title != sample_fields['study_title']:
            edited_sample_fields.append('study_title')
        sample_object.study_title = sample_fields['study_title']

    # sample_id
    if sample_fields['sample_id'] == '' or not sample_fields['sample_id']:
        sample_object.sample_id = None
    else:
        if sample_object.sample_id != sample_fields['sample_id']:
            edited_sample_fields.append('sample_id')
        sample_object.sample_id = sample_fields['sample_id']

    # species
    if sample_fields['species'] == '' or not sample_fields['species']:
        sample_object.species = None
    else:
        if sample_object.species != sample_fields['species']:
            edited_sample_fields.append('species')
        sample_object.species = sample_fields['species']

    # sample_matrix
    if sample_fields['sample_matrix'] == '' or not sample_fields['sample_matrix']:
        sample_object.sample_matrix = None
    else:
        if sample_object.sample_matrix != sample_fields['sample_matrix']:
            edited_sample_fields.append('sample_matrix')
        sample_object.sample_matrix = sample_fields['sample_matrix']

    # collection_protocol
    if sample_fields['collection_protocol'] == '' or not sample_fields['collection_protocol']:
        sample_object.collection_protocol = None
    else:
        if sample_object.collection_protocol != sample_fields['collection_protocol']:
            edited_sample_fields.append('collection_protocol')
        sample_object.collection_protocol = sample_fields['collection_protocol']

    if sample_fields['parent_type'] == '' or not sample_fields['parent_type']:
        sample_object.parent_type = None
    else:
        if sample_object.parent_type != sample_fields['parent_type']:
            edited_sample_fields.append('parent_type')
        sample_object.parent_type = sample_fields['parent_type']

    if sample_fields['parent_id'] == '' or not sample_fields['parent_id']:
        sample_object.parent_id = None
    else:
        if sample_object.parent_id != sample_fields['parent_id']:
            edited_sample_fields.append('parent_id')
        sample_object.parent_id = sample_fields['parent_id']

    if sample_fields['sample_storage_type'] == '' or not sample_fields['sample_storage_type']:
        sample_object.sample_storage_type = None
    else:
        if sample_object.sample_storage_type != sample_fields['sample_storage_type']:
            edited_sample_fields.append('sample_storage_type')
        sample_object.sample_storage_type = sample_fields['sample_storage_type']

    if sample_fields['hazard_group'] == '' or not sample_fields['hazard_group']:
        sample_object.hazard_group = None
    else:
        if sample_object.hazard_group != sample_fields['hazard_group']:
            edited_sample_fields.append('hazard_group')
        sample_object.hazard_group = sample_fields['hazard_group']

    if sample_fields['hazard_description'] == '' or not sample_fields['hazard_description']:
        sample_object.hazard_description = None
    else:
        if sample_object.hazard_description != sample_fields['hazard_description']:
            edited_sample_fields.append('hazard_description')
        sample_object.hazard_description = sample_fields['hazard_description']

    if sample_fields['campus'] == '' or not sample_fields['campus']:
        sample_object.campus = None
    else:
        if sample_object.campus != sample_fields['campus']:
            edited_sample_fields.append('campus')
        sample_object.campus = sample_fields['campus']

    if sample_fields['building'] == '' or not sample_fields['building']:
        sample_object.building = None
    else:
        if sample_object.building != sample_fields['building']:
            edited_sample_fields.append('building')
        sample_object.building = sample_fields['building']

    if sample_fields['room'] == '' or not sample_fields['room']:
        sample_object.room = None
    else:
        if sample_object.room != sample_fields['room']:
            edited_sample_fields.append('room')
        sample_object.room = sample_fields['room']

    if sample_fields['freezer_id'] == '' or not sample_fields['freezer_id']:
        sample_object.freezer_id = None
    else:
        if sample_object.freezer_id != sample_fields['freezer_id']:
            edited_sample_fields.append('freezer_id')
        sample_object.freezer_id = sample_fields['freezer_id']

    if sample_fields['shelf_id'] == '' or not sample_fields['shelf_id']:
        sample_object.shelf_id = None
    else:
        if sample_object.shelf_id != sample_fields['shelf_id']:
            edited_sample_fields.append('shelf_id')
        sample_object.shelf_id = sample_fields['shelf_id']

    if sample_fields['box_id'] == '' or not sample_fields['box_id']:
        sample_object.box_id = None
    else:
        if sample_object.box_id != sample_fields['box_id']:
            edited_sample_fields.append('box_id')
        sample_object.box_id = sample_fields['box_id']

    if sample_fields['tissue_bank_reference'] == '' or not sample_fields['tissue_bank_reference']:
        sample_object.tissue_bank_reference = None
    else:
        if sample_object.tissue_bank_reference != sample_fields['tissue_bank_reference']:
            edited_sample_fields.append('tissue_bank_reference')
        sample_object.tissue_bank_reference = sample_fields['tissue_bank_reference']

    return [sample_object,edited_sample_fields]



# Method for checking submitted changes.
def check_staging_samples(job_id,staging_id):

    job = start_job(job_id)

    # Get the staging object and commit it.
    staging_object = Staging.objects.filter(id=staging_id,status=1)[0]

    # Double encoded to include escape characters in view
    json_objects = json.loads(json.loads(staging_object.json))

    # Check the fields using model validators
    staging_object.field_validation = check_fields(json_objects,staging_object)

    staging_object.datetime_checked = now()

    val = json.loads(staging_object.field_validation)

    if len(val['failed_samples']) > 0:
        staging_object.status = 2

    else:
        staging_object.status = 3

    staging_object.save()

    # Update the job status
    finish_job(job)

def check_fields(json_objects,staging_object):

    passed_samples = []
    failed_samples = {}
    edited_fields = {}
    uniqueness_check = {}

    # Loop over the samples and validate them.
    for row_id,sample_fields in json_objects['objects'].items():

        error_map = {}

        # If its a number is likely to be a pk. Treat it as such.
        if not is_number(row_id):

            edited_fields[row_id] = []
            for field_name,value in sample_fields.items():
                edited_fields[row_id].append(field_name)

            sample_object = Sample(**sample_fields)
            sample_object.project = staging_object.project
            sample_object.last_edited_user = staging_object.user

            unique_key = sample_fields["sample_id"]+"-"+sample_fields["study_title"]

            if unique_key not in uniqueness_check:

                uniqueness_check[unique_key] = []

            uniqueness_check[unique_key].append(row_id)

            if(len(Sample.objects.filter(sample_id=sample_fields["sample_id"],
                                         study_title=sample_fields["study_title"])) > 0):
                error_map = {'sample_id':'Sample ID already exists!'}

        else:


            output = update_sample_model(row_id,sample_fields,staging_object.project)
            sample_object = output[0]

            sample_object.project = staging_object.project
            sample_object.last_edited_user = staging_object.user

            edited_fields[sample_object.id] = output[1]

            unique_key = sample_object.sample_id+"-"+sample_object.study_title

            if unique_key not in uniqueness_check:

                uniqueness_check[unique_key] = []

            uniqueness_check[unique_key].append(row_id)
            #uniqueness_check.append(sample_object.sample_id+"-"+sample_object.study_title)

        try:
            sample_object.full_clean()

        except Exception as e:

            for field,error in e:

                error_map[field] = error

            failed_samples[row_id] = error_map

        else:

            passed_samples.append(row_id)

    unique_check_results = check_uniques(failed_samples,passed_samples,uniqueness_check)
    passed_samples = unique_check_results[0]
    failed_samples = unique_check_results[1]


    return json.dumps({'passed_samples':passed_samples,
                        'failed_samples':failed_samples,
                        'edited_fields': edited_fields})

# Change this to use a dict with array as value. just loop over dict values, where count > 1, assign
def check_uniques(failed_samples,passed_samples,uniqueness_check):

   # print uniqueness_check

    for unique_id,entries in uniqueness_check.items():

        if len(entries) > 1:

            # Loop over the entries, if in passed, remove from passed
            # Add to failed with sample_id entry.
            for entry in entries:

                if entry in passed_samples:

                    # Remove from passed_samples
                    passed_samples.remove(entry)

                if entry not in failed_samples.keys():
                    failed_samples[entry] = {}

                failed_samples[entry]['sample_id'] = "Sample id is not unique in study"



            # Find the pks from the failed and passed samples
            # Remove them from the passed samples
            # Add them to the failed samples

            #pks = find_pks()

            #failed_samples[pk]['sample_id'] = 'Sample ID not unique in study!'

    return [passed_samples,failed_samples]

def validate_sample_file(job_id,staging_id,file_id):

    logger = logging.getLogger('django')

    job = start_job(job_id)

    # Get the staging object and commit it.
    staging_object = Staging.objects.filter(id=staging_id,status=1)[0]

    #file_object = File.objects.filter(id=file_id)[0]

    staging_object.file_validation = {'errors':[],'warnings':[]}

    # Open the file.
    # Check the column headers
    # Validate the fields
    wb = openpyxl.load_workbook(staging_object.file.filepath)

    # TODO:
    # Add the project to the staging_objects


    # Staging_

    #logger.error( json.dumps(wb.get_sheet_names()))
    #print "HERE--------------------->"
    #print wb.sheetnames[0]

    if len(wb.sheetnames) > 1:
        staging_object.file_validation['errors'].append('1 worksheet allowed per file')

    ws = wb[wb.sheetnames[0]]

    # Check the column headers
    #header = ws['A1:R1']
    validation = check_column_headers(ws['A1:T1'],staging_object.file_validation)

   # print validation
    file_validation = validation[0]
    header_map = validation[1]

    staging_object.file_validation = json.dumps(file_validation)
    # If no file errors, run the field validation
    if len(file_validation['errors']) == 0:

        json_output = build_json(ws,header_map,staging_object)

        #json_object = json_output[0]

        #sample_check_clash = json_output[1]

        #if sample_check_clash:
        #    file_validation['errors'].append('Edited samples have no PK field and multiple sample ids found in DB')


        staging_object.json = json.dumps(json.dumps(json_output))
        #staging_object.json = build_json(ws,header_map,staging_object)
        #staging_object.field_validation = check_file_fields(staging_object.json,header_map)
        #validation = check_fields(json_object)

        staging_object.field_validation = check_fields(json_output,staging_object)

        val = json.loads(staging_object.field_validation)

        if len(val['failed_samples']) > 0:
            staging_object.status = 2

        else:
            staging_object.status = 3

    else:
        staging_object.status = 2


    staging_object.datetime_checked = now()

    staging_object.save()

    job = finish_job(job)



def build_json(ws,header_map,staging_object):

    edited_pks = []
    new_pks = []
    edited_fields = {}
    objects = {}

    row_pos = 0
    for row in ws.iter_rows(min_row=2, max_col=20):

        sample_fields = {}
        edited_sample_fields = []

        if row[0].value == None:
            break

        for field_name, col_idx in header_map.items():

            # if (col_idx - 1) in row:
            #  # print "found"

            if col_idx and row[(col_idx - 1)]:

                sample_fields[field_name] = row[(col_idx - 1)].value

            # else:

            #   # print "not found"
        #print sample_fields


        # Lookup sample_id, study_title, and project.
        # If exist, append to edited_pks.
        # If not exist, append to new_pks

       # print sample_fields['sample_id']
       # print sample_fields['study_title']
       # print staging_object.project.id

        # Edited samples have the PKs entered
        if 'pk' in sample_fields and sample_fields['pk'] != '':
            existing_samples =  Sample.objects.filter(pk=sample_fields['pk'],
                                                    project=staging_object.project)

        else:
            existing_samples = []

        # Assume it is a dumb user who is editing entries without pks
      #  else:

       #     existing_samples = Sample.objects.filter(sample_id=sample_fields['sample_id'],
        #                                            study_title=sample_fields['study_title'],
         #                                           project=staging_object.project)

#        sample_check_clash = False

        # We found 1 entry... Hooray!
        if len(existing_samples) == 1:

           # print "found 1 ---------> \n"

            edited_pks.append(existing_samples[0].pk)

            objects[existing_samples[0].pk] = sample_fields

        # Uh-oh, there are multiple existing entries with these fields
#        elif len(existing_samples) > 1:
#
 #          # print "found >1 ---------> \n"

  #          sample_check_clash = True

        else:

           # print "found 0 ---------> \n"
            new_pks.append(sample_fields['sample_id']+'-'+sample_fields['study_title'])
            objects["new_"+str(row_pos)] = sample_fields

        row_pos = row_pos + 1

   # print(edited_pks)
   # print(new_pks)
   # print(objects)

    return {'edited_pks':edited_pks,
            'new_pks':new_pks,
            'objects':objects}

def get_default_import_header_map():

    return {'pk':None, 'study_title':None,'sample_id':None,'species':None,'sample_storage_type':None,
           'sample_matrix':None,'collection_protocol':None, 'campus':None,'building':None,
           'room':None,'freezer_id':None,'shelf_id':None,'box_id':None,'parent_type':None,
            'parent_id':None,'consent_form_information':None,'tissue_bank_reference':None,
            'hazard_group':None,'hazard_description':None}

def get_default_export_header_order():

    return [ 'pk','study_title','sample_id','species','sample_storage_type',
             'sample_matrix','collection_protocol', 'campus','building',
             'room','freezer_id','shelf_id','box_id','parent_type',
             'parent_id','consent_form_information','tissue_bank_reference',
             'hazard_group','hazard_description' ]

def check_column_headers(header,file_validation):

    # Loop over the headers and set the key value

    header_map = get_default_import_header_map()

    for col in header[0]:

        if col.value:

            col_value = col.value.strip().lower().replace(' ','_')

            #print col_value

            if col_value in header_map:

                header_map[col_value] = col.col_idx

            else:
                #file_validation['warnings'].append('Extra columns included')
                file_validation['warnings'].append('Extra columns included')
        #else:

         #   file_validation['errors'].append('Columns missing')



    for col,value in header_map.items():
        if value == None and col != 'pk':
            file_validation['errors'].append('Column missing: '+col)

    return [file_validation,header_map]

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False

def export_to_excel(job_id,user,sample_pks):

    job = start_job(job_id)

    for sample_pk in sample_pks:
       # print sample_pks
        sample_pk = int(sample_pk)

    samples = Sample.objects.filter(pk__in=sample_pks).order_by('pk')

    if len(samples) == 0:

        finish_job(job)
        return

    # Assume they are from the same project
    project_id = samples[0].project.pk

    # Create workbook
    wb = openpyxl.Workbook()

    # get worksheet
    ws = wb.active
    ws.title = "Samples from "+samples[0].project.title

    # get header
    header_order = get_default_export_header_order()

    # create header row
    for row in ws.iter_rows(min_row=1, max_col=len(header_order), max_row=1):
        p = 0
        for cell in row:
            cell.value = header_order[p]
            p = p + 1

    # create the other cells
    r = 0
    for row in ws.iter_rows(min_row=2, max_col=len(header_order), max_row=len(samples)+1):
        sample = samples[r]

        p = 0
        for cell in row:
            cell.value = getattr(sample,header_order[p])
            p = p + 1

        r = r + 1

    # Set filenames and save

    folder_path = os.path.join(default_storage.location,'user_uploads',str(project_id))

    if not os.path.isdir(folder_path):
        # # print os.path.join(default_storage.location,path_to_save)
        os.mkdir(folder_path)

    non_unique_name = "sampledb_sample_export_p:" + str(project_id) + "_u:" + user.username + "_" + now().strftime("%Y-%m-%d %H:%M:%S") + ".xlsx"

    file_name = build_and_check_file_name(folder_path,0,non_unique_name)

    full_path = os.path.join(folder_path,file_name)

    # Save the workbook
    wb.save(full_path)

    # Create the file
    file = File.objects.create(project=Project.objects.get(pk=project_id),
                               user_uploaded=user,
                               filename=file_name,
                               filepath=full_path,
                               datetime_uploaded=now(),
                               datetime_last_accessed = datetime.now(),
                               type='export')


    job = finish_job(job)

